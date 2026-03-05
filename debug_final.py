"""
CVS 10 SKU Scraper - just grab and print selector text directly
"""
import time, random
import pandas as pd
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

INPUT_FILE  = "CVS store brand skus.xlsx"
OUTPUT_FILE = "CVS_results.xlsx"
NUM_SKUS    = 10
CVS_BASE    = "https://www.cvs.com"

df   = pd.read_excel(INPUT_FILE, dtype=str).fillna("")
rows = df.head(NUM_SKUS)
print(f"SKUs: {list(rows['Sku'])}\n")

options = uc.ChromeOptions()
options.add_argument("--window-size=1280,800")
driver = uc.Chrome(options=options, headless=False, version_main=145)
driver.set_page_load_timeout(30)

def grab(css):
    """Return text from ALL matching elements, joined. Empty string if none found."""
    els = driver.find_elements(By.CSS_SELECTOR, css)
    return " ".join(e.text.strip() for e in els if e.text.strip())

results = []

try:
    for i, (_, row) in enumerate(rows.iterrows(), 1):
        sku  = str(row["Sku"]).strip()
        name = str(row["Cvs Item Description"]).strip()
        print(f"\n[{i}/{NUM_SKUS}] SKU: {sku} | {name[:60]}")

        result = {
            "Sku Nbr":             sku,
            "Product Name":        name,
            "Product URL":         "",
            "H/FSA Eligible":      "",
            "Product Description": "",
            "Ingredients":         "",
            "Scrape Status":       ""
        }

        # STEP 1: search → URL
        try:
            driver.get(f"{CVS_BASE}/search?searchTerm={sku}")
            tile = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "[id^='product-tile-link-']")
                )
            )
            href = tile.get_attribute("href")
            url  = href if href.startswith("http") else CVS_BASE + href
            result["Product URL"] = url
            print(f"  URL: {url}")

        except TimeoutException:
            body = driver.find_element(By.TAG_NAME, "body").text.lower()
            result["Scrape Status"] = (
                "ACCESS_DENIED" if "access denied" in body else
                "Not found"     if "sorry" in body else
                "Search timeout"
            )
            print(f"  ✗ {result['Scrape Status']}")
            results.append(result)
            pd.DataFrame(results).to_excel(OUTPUT_FILE, index=False)
            time.sleep(random.uniform(6, 10))
            continue

        # STEP 2: visit product page
        time.sleep(random.uniform(2, 4))
        try:
            driver.get(url)
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.TAG_NAME, "h1"))
            )
            time.sleep(2)
        except TimeoutException:
            body = driver.find_element(By.TAG_NAME, "body").text.lower()
            result["Scrape Status"] = (
                "ACCESS_DENIED on product page" if "access denied" in body
                else "Product page timeout"
            )
            print(f"  ✗ {result['Scrape Status']}")
            results.append(result)
            pd.DataFrame(results).to_excel(OUTPUT_FILE, index=False)
            time.sleep(random.uniform(6, 10))
            continue

        # STEP 3: just grab and store — no logic
        result["Product Description"] = grab("span.text-base.font-normal.not-italic.leading-6.text-black")
        result["H/FSA Eligible"]      = grab("#vendorDetailsBullet0 span")
        result["Ingredients"]         = grab("span.text-sm.font-normal.not-italic.leading-6.text-black.md\\:text-base") or "Ingredient not mentioned"

        print(f"  Description : {result['Product Description'][:80]}")
        print(f"  H/FSA       : {result['H/FSA Eligible']}")
        print(f"  Ingredients : {result['Ingredients'][:80]}")

        result["Scrape Status"] = "Success"
        results.append(result)
        pd.DataFrame(results).to_excel(OUTPUT_FILE, index=False)
        print(f"  Saved → {OUTPUT_FILE}")

        if i < NUM_SKUS:
            wait = random.uniform(6, 12)
            print(f"  Waiting {wait:.1f}s...")
            time.sleep(wait)

finally:
    driver.quit()

# Format Excel
wb = load_workbook(OUTPUT_FILE)
ws = wb.active
for cell in ws[1]:
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.font = Font(bold=True, color="FFFFFF")
for row in ws.iter_rows(min_row=2):
    sv = str(row[6].value or "")
    row[6].fill = PatternFill("solid", fgColor="C6EFCE" if sv == "Success" else "FFC7CE")
for col in ws.columns:
    w = max((len(str(c.value or "")) for c in col), default=10)
    ws.column_dimensions[col[0].column_letter].width = min(w + 4, 80)
wb.save(OUTPUT_FILE)

ok = sum(1 for r in results if r["Scrape Status"] == "Success")
print(f"\nDONE → {OUTPUT_FILE} | Success: {ok}/{len(results)}")