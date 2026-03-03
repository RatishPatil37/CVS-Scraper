import argparse, os, re, time, json
from dotenv import load_dotenv
load_dotenv()

import pandas as pd
from firecrawl import Firecrawl
from groq import Groq
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

CVS_BASE = "https://www.cvs.com"

OUTPUT_COLS = [
    "Sku Nbr", "Product Name", "Product URL",
    "H/FSA Eligible", "Product Description",
    "Ingredients", "Price", "UPC", "Brand", "Scrape Status",
]

# ── URL extraction ────────────────────────────────────────────────────────────

def _extract_all_product_hrefs(html):
    """Extract all /shop/.../prodid-XXXXXX hrefs from HTML."""
    pattern = re.compile(
        r'href=[\'"](/shop/[^\'"]*prodid-\d+[^\'"]*)[\'"]',
        re.IGNORECASE
    )
    seen = set()
    result = []
    for h in pattern.findall(html):
        if h not in seen:
            seen.add(h)
            result.append(h)
    return result


def _extract_main_results_hrefs(html):
    cutoff_patterns = [
        "recently-viewed", "Recently Viewed", "recentlyViewed", "recently_viewed",
        "sponsored-products", "Sponsored Products",
        "You might also like", "Related Products",
    ]
    trimmed = html
    lowest = len(html)
    for marker in cutoff_patterns:
        idx = trimmed.lower().find(marker.lower())
        if 0 < idx < lowest:
            lowest = idx
    trimmed = html[:lowest]
    return _extract_all_product_hrefs(trimmed)


def _best_match(hrefs, sku, upc=""):
    """Pick the href most likely matching this product."""
    if not hrefs:
        return None
    sku = str(sku).strip()
    upc = str(upc).strip() if upc else ""
    # Priority 1: prodid exactly matches SKU
    for h in hrefs:
        if re.search(rf'prodid-{re.escape(sku)}(\D|$)', h):
            return h
    # Priority 2: UPC appears in href
    if upc and len(upc) > 5:
        for h in hrefs:
            if upc in h:
                return h
    # Priority 3: first result on page
    return hrefs[0]


def _full_url(href):
    return CVS_BASE + href if href.startswith("/") else href


def _get_html(result):
    if isinstance(result, dict):
        return result.get("html") or result.get("rawHtml") or ""
    return getattr(result, "html", "") or getattr(result, "rawHtml", "") or ""


def _short_err(e):
    return str(e)[:150]


def _is_product_page(html, sku, upc=""):
    """Sanity check: does this HTML look like an actual product detail page?"""
    markers = ["add-to-cart", "AddToCart", "product-detail", "productDetail",
               "pdp-", "buybox", "buy-box", "add to cart"]
    has_structure = any(m.lower() in html.lower() for m in markers)
    has_sku = sku in html if sku else False
    has_upc = (upc in html) if (upc and len(upc) > 5) else False
    return has_structure or has_sku or has_upc


# ── 5-strategy URL resolution ─────────────────────────────────────────────────

def get_product_url(fc, sku, upc="", existing_url=""):
    """Returns (url, strategy_name) or (None, reason)."""
    sku = str(sku).strip()
    upc = str(upc).strip() if upc else ""

    # S0: use URL already in Excel
    if existing_url and "cvs.com/shop" in str(existing_url):
        print("    [S0] Using existing URL from Excel")
        return existing_url.strip(), "Excel URL"

    # S1: direct /shop/p/prodid-{SKU}
    direct = f"{CVS_BASE}/shop/p/prodid-{sku}"
    print(f"    [S1] Direct: {direct}")
    try:
        r = fc.scrape(direct, formats=["html"])
        html = _get_html(r)
        if html and _is_product_page(html, sku, upc):
            print("    [S1] ✓ Valid product page")
            return direct, "Direct prodid"
        print("    [S1] Page not a valid product page")
    except Exception as e:
        print(f"    [S1] ✗ {_short_err(e)}")

    # S2: search by SKU
    url = _try_search(fc, f"{CVS_BASE}/search?searchTerm={sku}", sku, upc, "S2 SKU search")
    if url:
        return url, "SKU search"

    # S3: search by UPC
    if upc and len(upc) > 5:
        url = _try_search(fc, f"{CVS_BASE}/search?searchTerm={upc}", sku, upc, "S3 UPC search")
        if url:
            return url, "UPC search"
    else:
        print("    [S3] Skipped (no UPC)")

    # S4: search with product type filter
    url = _try_search(fc, f"{CVS_BASE}/search?searchTerm={sku}&searchType=product", sku, upc, "S4 filtered search")
    if url:
        return url, "Filtered search"

    return None, "All strategies failed"


def _try_search(fc, search_url, sku, upc, label):
    print(f"    [{label}]: {search_url}")
    try:
        r = fc.scrape(search_url, formats=["html"])
        html = _get_html(r)
        if not html:
            print("      Empty HTML")
            return None
        hrefs = _extract_main_results_hrefs(html)
        print(f"      {len(hrefs)} links in main results")
        if not hrefs:
            hrefs = _extract_all_product_hrefs(html)
            print(f"      {len(hrefs)} links full-page fallback")
        if not hrefs:
            return None
        best = _best_match(hrefs, sku, upc)
        if best:
            url = _full_url(best)
            print(f"      ✓ {url}")
            return url
    except Exception as e:
        print(f"      Error: {_short_err(e)}")
    return None


# ── Scrape product page ───────────────────────────────────────────────────────

def scrape_product_page(fc, url):
    print(f"    -> Fetching markdown...")
    try:
        r = fc.scrape(url, formats=["markdown"])
        md = (r.get("markdown", "") if isinstance(r, dict)
              else getattr(r, "markdown", "") or "")
        if len(md) > 200:
            return md
        print(f"    -> Too short ({len(md)} chars)")
        return None
    except Exception as e:
        print(f"    -> Failed: {_short_err(e)}")
        return None


# ── Groq extraction ───────────────────────────────────────────────────────────

def extract_with_groq(groq_client, markdown, url):
    prompt = f"""Extract product info from this CVS.com product page (markdown format).

Return ONLY a raw JSON object (no fences, no explanation) with these exact keys:
{{
  "product_name": "full product name",
  "brand": "brand name only",
  "price": "$XX.XX or empty string",
  "upc": "numeric UPC if visible on page, else empty string",
  "hfsa_eligible": true or false,
  "product_description": "main description (2-5 sentences)",
  "ingredients": "full ingredients list as one string, or empty string"
}}

RULES:
- hfsa_eligible = true if you see: FSA, HSA, H/FSA, FSA Eligible, HSA Eligible, Health Savings Account, Flexible Spending
- Return ONLY raw JSON, nothing else

Product URL: {url}

Page:
{markdown[:7000]}
"""
    try:
        resp = groq_client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": prompt}],
            temperature=0, max_tokens=1500,
        )
        text = resp.choices[0].message.content.strip()
        text = re.sub(r"^```json\s*|^```\s*|```$", "", text, flags=re.MULTILINE).strip()
        return json.loads(text)
    except json.JSONDecodeError as e:
        print(f"    -> JSON parse error: {e}")
        return {}
    except Exception as e:
        print(f"    -> Groq error: {e}")
        return {}


# ── Per-SKU orchestration ─────────────────────────────────────────────────────

def scrape_sku(fc, groq_client, sku, name, upc="", existing_url=""):
    row = {col: "" for col in OUTPUT_COLS}
    row["Sku Nbr"] = sku
    row["Product Name"] = name
    row["UPC"] = upc

    url, strategy = get_product_url(fc, sku, upc, existing_url)
    if not url:
        row["Scrape Status"] = "URL not found"
        return row

    row["Product URL"] = url
    md = scrape_product_page(fc, url)
    if not md:
        row["Scrape Status"] = "Page fetch failed"
        return row

    data = extract_with_groq(groq_client, md, url)
    if not data:
        row["Scrape Status"] = "Groq extraction failed"
        return row

    row["Product Name"]        = data.get("product_name") or name
    row["Brand"]               = data.get("brand", "")
    row["Price"]               = data.get("price", "")
    row["UPC"]                 = data.get("upc") or upc
    row["H/FSA Eligible"]      = "Yes" if data.get("hfsa_eligible") else "No"
    row["Product Description"] = data.get("product_description", "")
    row["Ingredients"]         = data.get("ingredients", "")
    row["Scrape Status"]       = f"Success [{strategy}]"
    return row


# ── Save Excel ────────────────────────────────────────────────────────────────

def save_excel(results, path):
    pd.DataFrame(results, columns=OUTPUT_COLS).to_excel(path, index=False)
    wb = load_workbook(path)
    ws = wb.active
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.font = Font(bold=True, color="FFFFFF")
    hi = OUTPUT_COLS.index("H/FSA Eligible")
    si = OUTPUT_COLS.index("Scrape Status")
    green  = PatternFill("solid", fgColor="C6EFCE")
    red    = PatternFill("solid", fgColor="FFC7CE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    for row in ws.iter_rows(min_row=2):
        sv = str(row[si].value or "")
        if str(row[hi].value or "") == "Yes":
            row[hi].fill = green
        if sv.startswith("Success"):
            row[si].fill = green
        elif "failed" in sv.lower() or "not found" in sv.lower():
            row[si].fill = red
        else:
            row[si].fill = yellow
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 80)
    wb.save(path)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input",  default="CVS store brand skus.xlsx")
    ap.add_argument("--output", default="CVS_results_v2.xlsx")
    ap.add_argument("--limit",  type=int,   default=None)
    ap.add_argument("--delay",  type=float, default=2.5)
    args = ap.parse_args()

    fc_key   = os.environ.get("FIRECRAWL_API_KEY") or exit("Missing FIRECRAWL_API_KEY")
    groq_key = os.environ.get("GROQ_API_KEY")      or exit("Missing GROQ_API_KEY")
    fc = Firecrawl(api_key=fc_key)
    groq_client = Groq(api_key=groq_key)

    print(f"\nReading {args.input}...")
    df = pd.read_excel(args.input, dtype=str).fillna("")
    print(f"Columns: {list(df.columns)}")

    cols_lower = {c.lower(): c for c in df.columns}
    def find_col(*kws):
        for kw in kws:
            for lo, orig in cols_lower.items():
                if kw in lo: return orig
        return None

    sku_col  = find_col("sku")          or df.columns[0]
    name_col = find_col("desc", "name") or df.columns[1]
    upc_col  = find_col("upc")
    url_col  = find_col("product url", "url")

    print(f"SKU col: '{sku_col}' | Name col: '{name_col}' | UPC col: '{upc_col}' | URL col: '{url_col}'")

    rows = df[df[sku_col].str.strip() != ""]
    if args.limit:
        rows = rows.head(args.limit)

    total = len(rows)
    print(f"\nProcessing {total} SKUs...\n" + "="*70)

    results = []
    for i, (_, row) in enumerate(rows.iterrows(), 1):
        sku  = str(row[sku_col]).strip()
        name = str(row[name_col]).strip()
        upc  = str(row[upc_col]).strip() if upc_col else ""
        eurl = str(row[url_col]).strip() if url_col else ""

        print(f"\n[{i}/{total}] SKU: {sku}  |  {name[:60]}")
        if upc: print(f"  UPC: {upc}")

        result = scrape_sku(fc, groq_client, sku, name, upc, eurl)
        results.append(result)

        icon = "✓" if result["Scrape Status"].startswith("Success") else "✗"
        print(f"  {icon} {result['Scrape Status']} | H/FSA: {result['H/FSA Eligible']}")
        print("-"*70)

        if i % 5 == 0:
            save_excel(results, args.output)
            print(f"  [Auto-saved {i} rows]")

        time.sleep(args.delay)

    save_excel(results, args.output)

    ok   = sum(1 for r in results if r["Scrape Status"].startswith("Success"))
    hfsa = sum(1 for r in results if r["H/FSA Eligible"] == "Yes")
    print(f"\n{'='*70}\nDONE → {args.output}")
    print(f"Total: {total} | Success: {ok} ({ok/total*100:.0f}%) | Failed: {total-ok} | H/FSA: {hfsa}")

    fails = [r for r in results if not r["Scrape Status"].startswith("Success")]
    if fails:
        print("\nFailed SKUs:")
        for r in fails:
            print(f"  SKU {r['Sku Nbr']:>10} | {r['Product Name'][:50]} | {r['Scrape Status']}")


if __name__ == "__main__":
    main()
